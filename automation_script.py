import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook

# Setting up the Selenium WebDriver with Service
driver_path = './chromedriver.exe'  # Use relative path if chromedriver is in the same directory
service = Service(driver_path)
options = Options()
options.add_argument("--start-maximized")  # Optionally start browser maximized

# Initialize the WebDriver with the Service and Options
driver = webdriver.Chrome(service=service, options=options)

# Function to search Google and extract the longest and shortest options
def get_search_suggestions(keyword):
    driver.get('https://www.google.com')
    search_box = driver.find_element(By.NAME, 'q')
    search_box.send_keys(keyword)
    time.sleep(2)  # Wait for suggestions to load

    # Get all the suggestion elements
    suggestions = driver.find_elements(By.CSS_SELECTOR, 'li span')

    longest = ""
    shortest = ""
    for suggestion in suggestions:
        text = suggestion.text
        if text:
            if not longest or len(text) > len(longest):
                longest = text
            if not shortest or len(text) < len(shortest):
                shortest = text

    return longest, shortest

# Function to process Excel file based on the current day of the week
def process_excel():
    today = datetime.datetime.now().strftime('%A')  # Get the current day of the week
    print(f"Today is: {today}")
    
    # Load the workbook and select the first sheet
    wb = load_workbook('QUPS_keywords.xlsx')  # Ensure this is the correct file name
    sheet = wb.active

    # Print the column headers for debugging
    print([col[0].value for col in sheet.iter_cols(1, sheet.max_column)])

    # Find the column for the current day of the week
    day_column = None
    for col in sheet.iter_cols(1, sheet.max_column):
        if col[0].value.strip().lower() == today.lower():  # Match day with case-insensitive comparison
            day_column = col
            break

    if day_column is None:
        print(f"No data for today ({today}) in the Excel file.")
        return

    # Iterate through each keyword in the selected column and process it
    for row in range(1, len(day_column)):
        keyword = day_column[row].value
        if keyword:
            longest, shortest = get_search_suggestions(keyword)
            # Write the results back into the Excel file (Columns B and C for longest and shortest)
            sheet[f'B{row+1}'] = longest  # Write longest option in column B
            sheet[f'C{row+1}'] = shortest  # Write shortest option in column C

    # Save the modified Excel file with the updated results
    wb.save('modified_QUPS_keywords.xlsx')  # Save as a new file or overwrite the original

# Run the script
try:
    process_excel()
finally:
    driver.quit()
